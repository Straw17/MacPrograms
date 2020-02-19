#Imports:
from openpyxl import Workbook
from openpyxl import load_workbook

#Variable setup 1:
wb = load_workbook('STV.xlsx') #Replace with file name.
ws = wb.active #Sets active sheet.
votes = {} #Votes for each person.
voters = {} #Voters for each person.
voterNomLen = {} #Number of candidates each person has nominated.
voteValue = {} #Value of each person's vote.
votersList = [] #List of voters.
voterNum = 0 #Number of voters.
minVotes = 0 #Minimum number of votes required to win.
cansList = [] #List of candidates.
invalVoters = [] #Removed voters.
invalCans = [] #Removed candidates.
winners = [] #List of winners.
letterList = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
              'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
              'U', 'V', 'W', 'X', 'Y', 'Z',
              'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ',
              'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT',
              'AU', 'AV', 'AW', 'AX', 'AY', 'AZ'] #List of X values in spreadsheet.

#Data collection functions:
def votersListFunction(letterList, ws, votersList):
    #Collects a list of voters
    global voteValue
    for i in range(50):
        cell = ws[str(letterList[i])+'1']
        if cell.value == None:
            break
        votersList.append(str(letterList[i]))
        voteValue.setdefault(letterList[i], 1)
    return votersList

#Variable setup 2:
votersList = votersListFunction(letterList, ws, votersList)

#Base internal functions:
def voteCollection(votes2, voters2, votersList2, cansList2, invalVoters2, ws, winners, voteValue2, invalCans):
    #Collects votes.
    voters2 = {}
    votes2 = {}
    cansList2 = []
    nextInvalVoters2 = []
    TBR = []
    for a in range(len(votersList2)):
        b = 1
        c = 1
        check = 1
        while True:
            cell = ws[votersList2[a] + str(b)]
            if cell.value == None:
                TBR.append(votersList2[a])
                break
            if cell.value in winners or cell.value in invalCans:
                b += 1
            else:
                if cell.value not in cansList2:
                    cansList2.append(cell.value)
                votes2.setdefault(cell.value, 0)
                votes2[cell.value] += voteValue[votersList2[a]]
                voters2.setdefault(cell.value, [])
                voters2[cell.value].append(votersList2[a])
                break
    for i in range(len(TBR)):
        invalVoters2.append(TBR[i])
        del voteValue2[TBR[i]]
        votersList2.remove(TBR[i])
    global votes
    global voters
    global votersList
    global invalVoters
    global cansList
    votes = votes2
    voters = voters2
    votersList = votersList2
    invalVoters = invalVoters2
    cansList = cansList2

def winCheck(votes, votersList, voters, cansList):
    #Checks for wins and loses.
    global winners
    global invalCans
    global voteValue
    global minVotes
    minVotes = len(votersList)/3
    for i in range(len(votes)):
        if votes[cansList[i]] >= minVotes:
            winners.append(cansList[i])
            winVoters = voters[cansList[i]]
            surplus = (votes[cansList[i]] - minVotes)/len(voters[cansList[i]])
            for x in range(len(winVoters)):
                voteValue[winVoters[x]] = surplus
            return True
    invalCan = (min(votes, key = votes.get))
    invalCans.append(invalCan)
    invalCanVoters = voters[invalCan]
    return False

#Run program:
while len(winners) != 3:
    voteCollection(votes, voters, votersList, cansList, invalVoters, ws, winners, voteValue, invalCans)
    winCheck(votes, votersList, voters, cansList)
print(winners)
